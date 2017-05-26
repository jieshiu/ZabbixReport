#!/usr/bin/python
# -*- coding: utf-8 -*-

import re
import time
from datetime import datetime
import MySQLdb

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell

from openpyxl.shared.exc import CellCoordinatesException, \
        SheetTitleException, InsufficientCoordinatesException, \
        NamedRangeException
from openpyxl.writer.worksheet import write_worksheet
from drawgraph import drawpic
import Image
import timeit

Aserver_ip_list=[]

class zabbixAvgItem:  		
  hostid=0
  intfrom=0
  inttill=0  
  strkey=''
  strtablename=''
  objcursor= None
  avgvalue=0
  maxvalue=0
  minvalue=0    
  strsql=''   
  def __init__(self,o,k,n,h,f,t,m):  
      self.hostid = h  
      self.intfrom = f  
      self.inttill = t  
      self.strkey=k
      self.objcursor=o
      self.strtablename=n
      self.mode=m
      self.graphpath=''
      self.graphpath2=''
      self.strsql="select b.value,b.clock from items a,%s b where a.itemid=b.itemid and a.key_ ='%s' and a.hostid=%d and b.clock > %d and b.clock < %d"%(self.strtablename,self.strkey,self.hostid,self.intfrom,self.inttill)
      if self.mode=='trends':
      	 self.strsql="select b.value_min,b.value_avg,b.value_max,b.clock from items a,%s b where a.itemid=b.itemid and a.key_ ='%s' and a.hostid=%d and b.clock > %d and b.clock < %d"%(self.strtablename,self.strkey,self.hostid,self.intfrom,self.inttill)            
  def __del__(self):
      pass				
  def runcount(self,ipname,statname,tagname,stime,outgraph,outpath,fontpath):
      self.avgvalue = 0
      self.maxvalue = 0
      self.minvalue = 100
      self.graphpath='%s%s_%s'%(outpath,ipname,tagname)+'.png'
      self.graphpath2='%s%s_%s'%(outpath,ipname,tagname)+'.jpg'
      #print self.__strsql
      data_list=[]
      dates_list=[]
      #print self.strsql
      self.objcursor.execute(self.strsql)        		
      numrows = int(self.objcursor.rowcount)    			
      if numrows>0:	    		
        for i in range(numrows):	        		
           row = self.objcursor.fetchone() 
           #if self.strkey== 'perf_counter[\\\\PhysicalDisk(_Total)\\\\% Idle Time]':          
              #print row['value_min']    
           dates_list.append(datetime.fromtimestamp(row['clock']))
           if self.mode=='trends':
              data_list.append(row['value_avg']) 
              self.avgvalue=self.avgvalue + row['value_avg']
              if self.maxvalue < row['value_max']:
                 self.maxvalue=row['value_max']
              if self.minvalue > row['value_min']:
                 if row['value_min']>0:
                    self.minvalue=row['value_min']
           else:     
              data_list.append(row['value']) 
              self.avgvalue=self.avgvalue + row['value']
              if self.maxvalue < row['value']:
                 self.maxvalue=row['value']
              if self.minvalue > row['value']:
                 if row['value']>0:
                    self.minvalue=row['value']
        self.avgvalue=self.avgvalue/numrows           
        if outgraph == 'yes':
           #print 'build graph...wait'
           p=drawpic(data=data_list,name=self.graphpath,dates=dates_list,fontp=fontpath)
           p._Gpic('date','traffic',u'%s年%s月%s日%s_%s趋势图' %(stime.split('-')[0],stime.split('-')[1],stime.split('-')[2],ipname,statname))
           #print 'create graph...end'
           Image.open(self.graphpath).save(self.graphpath2,'JPEG')
        return 1 	
      else:
        return 0
  def getavgvalue(self):      			  		
      return self.avgvalue
  def getmaxvalue(self):      			  		
      return self.maxvalue
  def getminvalue(self):
      if self.minvalue==100:
    	 return 0
      return self.minvalue
  def getgraphpath(self):
  	  return self.graphpath2

def formatws(wbsheet,strtime):
    wbsheet.cell('A1').value = '%s'%('服务器名称')
    wbsheet.cell('B1').value = '%s'%('设备IP')
    wbsheet.cell('C2').value = '%s'%('cpu使用率均值')
    wbsheet.cell('D2').value = '%s'%('cpu使用率峰值')
    wbsheet.cell('E2').value = '%s'%('内存空闲均值GB')
    wbsheet.cell('F2').value = '%s'%('最小内存空闲GB')
    wbsheet.cell('G2').value = '%s'%('磁盘idle均值')
    wbsheet.cell('H2').value = '%s'%('最小磁盘idle')
    #wbsheet.cell('C1').value = '%s'%(strtime)
    #wbsheet.range(wbsheet.cell('C1'), wbsheet.cell('H1')).value= '%s'%(strtime)    
    #wbsheet.cell('B1').style.font.color.index =Color.GREEN   
    #wbsheet.cell('B1').style.font.name ='Arial'   
    #wbsheet.cell('B1').style.font.size =8

    wbsheet.column_dimensions[wbsheet.cell('A1').column].width =18.0
    wbsheet.cell('A1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('A1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('A1').style.font.bold =True
    wbsheet.cell('A1').style.alignment.wrap_text =True
    wbsheet.cell('A2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('A2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('A2').style.font.bold =True
    wbsheet.cell('A2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('B1').column].width =18.0
    wbsheet.cell('B1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('B1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('B1').style.font.bold =True
    wbsheet.cell('B1').style.alignment.wrap_text =True
    wbsheet.cell('B2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('B2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('B2').style.font.bold =True
    wbsheet.cell('B2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('C1').column].width =18.0
    wbsheet.cell('C1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('C1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('C1').style.font.bold =True
    wbsheet.cell('C1').style.alignment.wrap_text =True
    wbsheet.cell('C2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('C2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('C2').style.font.bold =True
    wbsheet.cell('C2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('D1').column].width =18.0
    wbsheet.cell('D1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('D1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('D1').style.font.bold =True
    wbsheet.cell('D1').style.alignment.wrap_text =True
    wbsheet.cell('D2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('D2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('D2').style.font.bold =True
    wbsheet.cell('D2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('E1').column].width =18.0
    wbsheet.cell('E1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('E1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('E1').style.font.bold =True
    wbsheet.cell('E1').style.alignment.wrap_text =True
    wbsheet.cell('E2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('E2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('E2').style.font.bold =True
    wbsheet.cell('E2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('F1').column].width =18.0
    wbsheet.cell('F1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('F1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('F1').style.font.bold =True
    wbsheet.cell('F1').style.alignment.wrap_text =True
    wbsheet.cell('F2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('F2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('F2').style.font.bold =True
    wbsheet.cell('F2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('G1').column].width =18.0
    wbsheet.cell('G1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('G1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('G1').style.font.bold =True
    wbsheet.cell('G1').style.alignment.wrap_text =True
    wbsheet.cell('G2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('G2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('G2').style.font.bold =True
    wbsheet.cell('G2').style.alignment.wrap_text =True
    wbsheet.column_dimensions[wbsheet.cell('H1').column].width =18.0
    wbsheet.cell('H1').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('H1').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('H1').style.font.bold =True
    wbsheet.cell('H1').style.alignment.wrap_text =True
    wbsheet.cell('H2').style.fill.fill_type =Fill.FILL_SOLID
    wbsheet.cell('H2').style.fill.start_color.index ='C5D9F1'
    wbsheet.cell('H2').style.font.bold =True
    wbsheet.cell('H2').style.alignment.wrap_text =True

    #wbsheet.row_dimensions[wbsheet.cell('F1').row].height = 30   
    return

def isaserver(ipaddre):
    for item in Aserver_ip_list:      	    
        if ipaddre == item:        	
           return 1
    return 0
    
def getexceldata(ws,cursor,cursor2,intfrom,inttill,outgraph,outpath,fontpath):    
    index_num=0
    row_num=2
    numrows=0
    pddata={'table':{},'pic':{}}
    mode='history'
    time_tuple = time.localtime(intfrom)
    date_str = time.strftime("%Y-%m-%d", time_tuple)
    dt_start = datetime.fromtimestamp(intfrom)
    days=(datetime.now()-dt_start).days    
    if days > 7:
    	 mode='trends'    
    sql = 'select a.hostid,a.name,b.ip from hosts a,interface b where a.hostid=b.hostid and a.status=0'    
    cursor.execute(sql)
    alldata = cursor.fetchall()
    if alldata:  
       for rec in alldata:          	       
          if isaserver(rec['ip']) == 0:
          	 continue
          print rec['name'],rec['ip']
          pddata['table'].setdefault(rec['ip'],[])          
          ws.cell(row=row_num,column=0).value = rec['name']
          ws.cell(row=row_num,column=1).value = rec['ip']
    
          #cpu使用率均值          
          keyname=u'cpu使用率均值'
          p=zabbixAvgItem(cursor2,'perf_counter[\\\\Processor(_Total)\\\\% Processor Time]',mode,rec['hostid'],intfrom,inttill,mode)  
          p.runcount(rec['ip'],keyname,'cpuavg',date_str,outgraph,outpath,fontpath)  
          exstr="%.3f %%" %(p.getavgvalue())  
          #print exstr
          ws.cell(row=row_num,column=2).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath())          
              
          #cpu使用率峰值
          keyname=u'cpu使用率峰值'
          p=zabbixAvgItem(cursor2,'perf_counter[\\\\Processor(_Total)\\\\% Processor Time]',mode,rec['hostid'],intfrom,inttill,mode)    
          p.runcount(rec['ip'],keyname,'cpumax',date_str,outgraph,outpath,fontpath)
          exstr="%.3f %%" %(p.getmaxvalue())  
          #print exstr
          ws.cell(row=row_num,column=3).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath()) 
    
          #内存空闲均值GB   
          keyname=u'内存空闲均值'       
          p=zabbixAvgItem(cursor2,'vm.memory.size[free]',mode+'_uint',rec['hostid'],intfrom,inttill,mode) 
          p.runcount(rec['ip'],keyname,'memavg',date_str,outgraph,outpath,fontpath)   
          exstr="%d" %(p.getavgvalue()/1024/1024/1024)
          #print exstr
          ws.cell(row=row_num,column=4).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath()) 
    
	  #最小内存空闲GB 
	  keyname=u'最小内存空闲'
          p=zabbixAvgItem(cursor2,'vm.memory.size[free]',mode+'_uint',rec['hostid'],intfrom,inttill,mode) 
          p.runcount(rec['ip'],keyname,'memmin',date_str,outgraph,outpath,fontpath)   
          exstr="%d" %(p.getmaxvalue()/1024/1024/1024)
          #print exstr
          ws.cell(row=row_num,column=5).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath()) 
	  
	  #磁盘idle均值      
	  keyname=u'磁盘idle均值'    
          p=zabbixAvgItem(cursor2,'perf_counter[\\\\PhysicalDisk(_Total)\\\\% Idle Time]',mode,rec['hostid'],intfrom,inttill,mode) 
          p.runcount(rec['ip'],keyname,'diskavgidle',date_str,outgraph,outpath,fontpath)  
          exstr="%.3f %%" %(p.getavgvalue())   
          #print exstr
          ws.cell(row=row_num,column=6).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath()) 
	  
	  #最小磁盘idle
	  keyname=u'最小磁盘idle'  
          p=zabbixAvgItem(cursor2,'perf_counter[\\\\PhysicalDisk(_Total)\\\\% Idle Time]',mode,rec['hostid'],intfrom,inttill,mode) 
          p.runcount(rec['ip'],keyname,'diskminidle',date_str,outgraph,outpath,fontpath)
          exstr="%.3f %%" %(p.getminvalue())  
          #print exstr
          ws.cell(row=row_num,column=7).value = exstr
          pddata['table'][rec['ip']].append(exstr)
          pddata['pic'].setdefault(rec['name']+'_'+keyname,p.getgraphpath()) 
	    
          row_num+=1
          index_num+=1             
    return pddata

def genexcel(**kwargs):
    global Aserver_ip_list
    startclock = time.clock()
    iplist=''
    outpath=''
    excel_filename=''
    stattime=''
    hostip=''
    usrname=''
    usrpsw=''
    dbname=''
    outgraph=''
    fontpath=''
    params = kwargs.keys()
    #params.sort()
    for kw in params:
        if kw=='outpath':
	   outpath=kwargs[kw]
	if kw=='excel_filename':
	   excel_filename=kwargs[kw]
	if kw=='stattime':
	   stattime=kwargs[kw]
  	if kw=='hostip':
	   hostip=kwargs[kw]
  	if kw=='usrname':
	   usrname=kwargs[kw]
	if kw=='usrpsw':
	   usrpsw=kwargs[kw]
	if kw=='dbname':
	   dbname=kwargs[kw]
	if kw=='outgraph':
	   outgraph=kwargs[kw]	  
	if kw=='fontpath':
	   fontpath=kwargs[kw]	
	if kw=='iplist':
	   iplist=kwargs[kw]
	  	
#def genexcel(outpath,excel_filename, stattime, hostip,usrname,usrpsw,dbname):      
    Aserver_ip_list=iplist.split(',')
    print Aserver_ip_list
    
    try: 
        conn = MySQLdb.connect(host=hostip,user=usrname,passwd=usrpsw,db=dbname,charset='utf8')
    except Exception, e: 
        print e 
        sys.exit() 

    cursor = conn.cursor(cursorclass = MySQLdb.cursors.DictCursor)
    cursor2 = conn.cursor(cursorclass = MySQLdb.cursors.DictCursor)
    localtm=time.localtime(time.time())
    if stattime != '':
       stattimelist=stattime.split(',')
       fromtime = time.strptime(stattimelist[0], '%Y%m%d_%H_%M_%S')
       tilltime = time.strptime(stattimelist[1], '%Y%m%d_%H_%M_%S')       
       fromlocal=time.localtime(time.mktime(fromtime))   
       tilllocal=time.localtime(time.mktime(tilltime)) 	 
    else:
       print 'stattime is need!' 
       sys.exit() 
    strfrom = time.strftime(u'%Y-%m-%d %H:%M:%S',fromlocal)
    strtill = time.strftime(u'%Y-%m-%d %H:%M:%S',tilllocal)
    strtitle = time.strftime(u'%m月%d日',fromlocal)
    strhead = time.strftime(u'%m月%d日%A',fromlocal)
    # 时间戳
    print strfrom
    print strtill
    intfrom=int(time.mktime(time.strptime(strfrom,'%Y-%m-%d %H:%M:%S')))
    inttill=int(time.mktime(time.strptime(strtill,'%Y-%m-%d %H:%M:%S')))

    wb = Workbook()
    #ws = wb.create_sheet(0)  # insert at the first position
    ws = wb.worksheets[0]
    ws.title = strtitle   
    formatws(ws,strhead)
    
    outdata=getexceldata(ws,cursor,cursor2,intfrom,inttill,outgraph,outpath,fontpath)
    
    #for host in cursor.fetchall():  
    #  ws.cell(row=row_num,column=0)=host 
    #  row_num+=1
    cursor.close()
    cursor2.close()
    #conn.commit()
    conn.close()

    #file_name = 'test.xlsx'    
    dest_filename = '%s%s'%(outpath,excel_filename) 
    #ew = ExcelWriter(workbook = wb) 
    #ew = ExcelWriter(workbook = wb)
    #dest_filename = r'empty_book.xlsx'
    elapsed = (time.clock() - startclock)
    print("Time used:",elapsed)
    wb.save(filename = dest_filename)    
    return outdata
